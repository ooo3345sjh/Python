"""
날짜 : 2023/01/12
이름 : 서정현
내용 : 파이썬 외부 패키지 모듈 실습
"""

from openpyxl import Workbook, load_workbook

# 새로운 엑셀파일 생성
workbook = Workbook()

# 첫번째 시트 활성화
sheet = workbook.active

# 데이터 입력
sheet['A1'] = '파이썬 엑셀 실습'
sheet.append(['아이디', '이름', '전화번호', '나이', '주소'])
sheet.append(['a101', '김유신', '010-1234-1001', 25, '김해시'])
sheet.append(['a102', '김춘추', '010-1234-1002', 22, '경주시'])
sheet.append(['a103', '장보고', '010-1234-1003', 35, '완도시'])
sheet.append(['a104', '강감찬', '010-1234-1004', 45, '경기도'])
sheet.append(['a105', '이순신', '010-1234-1005', 55, '서울시'])

# 저장닫기
workbook.save('C:/Users/java2/Desktop/Member.xlsx')
workbook.close()

print('프로그램 종료...')


# 엑셀 파일 읽기
# data_only=Ture로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("C:/Users/java2/Desktop/Member.xlsx", data_only=True)

#시트 이름으로 불러오기
load_ws = load_wb['Sheet']

#셀 주소로 값 출력
print(load_ws['A1'].value)

get_cells = load_ws['A1':'E7']

for row in get_cells:
        for cell in row:
            if cell.value != None:
                print('%s ' % cell.value, end='')
        print()