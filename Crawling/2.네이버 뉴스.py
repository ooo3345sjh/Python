"""
날짜 : 2023/01/16
이름 : 서정현
내용 : 파이썬 네이버 뉴스 크롤링 실습하기
"""
# 외부라이브러리 pip install requests
# 외부라이브러리 pip install bs4
import requests as req
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, load_workbook

pg = 1
count = 1
Prelists = ""

# 새로운 엑셀파일 생성
workbook = Workbook()

# 첫번째 시트 활성화
sheet = workbook.active

# 데이터 입력
sheet['A1'] = '네이버 뉴스 크롤링 실습'
sheet.merge_cells('A1:C1') # 셀병합

while True:

    # HTML 요청
    url = 'https://news.naver.com/main/list.naver?mode=LS2D&sid2=230&sid1=105&mid=shm&page=%d' % pg
    html = req.get(url, headers={'User-Agent':'Mozilla/5.0'}).text
    # print(html)

    # 문서 객체 생성
    dom = bs(html, 'html.parser')

    # 데이터 파싱
    tit = dom.select_one('#main_content > div.list_header.newsflash_header > h3').text
    # print('tit :', tit)

    lists = dom.select("#main_content > div.list_body.newsflash_body > ul> li")

    # pg값이 다르면 종료
    page_tag = dom.select_one('#main_content > div.paging > strong').text
    if int(pg) != int(page_tag): break

    for list in lists:
        tag_a = list.select_one('dl > dt:not(.photo) > a')
        title = tag_a.text
        href = tag_a['href']
        # print('count :', count)
        # print('title :', title.strip())
        # print('href :', href.strip())

        # 데이터 입력
        sheet.append([count, title, href])
        print('%d건...', count)

        count += 1

    pg += 1

# 저장닫기
workbook.save('C:/Users/java2/Desktop/naverNews.xlsx')
workbook.close()
print('프로그램 종료...') 





