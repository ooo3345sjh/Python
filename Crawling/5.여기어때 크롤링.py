"""
날짜 : 2023/03/03
이름 : 서정현
내용 : 여기어때 데이터 크롤링
"""
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
import pymysql
import time
from urllib.request import urlretrieve
import uuid
import os
import requests, json
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, load_workbook
from random import randrange

# 태그 제거 방법
import re

def remove_html(sentence) :
	sentence = re.sub('(<([^>]+)>)', '', sentence)
	return sentence

# 카카오맵 API 요청 및 지오코딩 함수 임포트
def get_location(address):
    url = 'https://dapi.kakao.com/v2/local/search/address.json?query=' + address
    # 'KaKaoAK '는 그대로 두고 개인키만 지우고 입력
    headers = {"Authorization": "KakaoAK 774c9bcb3bb773884b71317afda40365"}
    api_json = json.loads(requests.get(url,headers=headers).text)
    try:
        address = api_json['documents'][0]['road_address']
        lat = address['y'] # 위도
        lng = address['x'] # 경도
        zone_no = address['zone_no'] # 우편번호
        address_name = address['address_name'] # 기본 주소
        building_name = address['building_name'] # 상세 주소
        region_2depth_name = address['region_2depth_name'] # 군, 구
        return [lat, lng, zone_no, address_name, building_name, region_2depth_name]
    except:
        return [None, None, None, None, None, None]
    


# 전국8도 및 광역시 번호 
def get_province_no(address):
    if  "강원" in address:
        return "1"
    elif "경기" in address:
        return "2"
    elif "경남" in address:
        return "3"
    elif "경북" in address:
        return "4"
    elif "광주" in address:
        return "5"
    elif "대구" in address:
        return "6"
    elif "대전" in address:
        return "7"
    elif "부산" in address:
        return "8"
    elif "서울" in address:
        return "9"
    elif "울산" in address:
        return "10"
    elif "인천" in address:
        return "11"
    elif "전남" in address:
        return "12"
    elif "전북" in address:
        return "13"
    elif "제주" in address:
        return "14"
    elif "충남" in address:
        return "15"
    elif "충북" in address:
        return "16"
    else:
        return "17"

# 숙소유형
def get_accType_no(cate):
    if cate == 1:
        return "1"
    elif cate == 2:
        return "2"
    elif cate == 3:
        return "3"
    elif cate == 6:
        return "4"
    elif cate == 5:
        return "5"

# 이미지 저장 후 이미지 이름들을 문자열로 합쳐서 반환
def get_imgArrStr(images, cate, acc_name):
    
    # 이미지를 저장할 폴더
    img_directory = 'C:\\Users\\ooo33.DESKTOP-56U45AS\\Desktop\\img\\{}'.format(cate)
        
    # 폴더가 없으면 새로 생성
    if not os.path.isdir(img_directory):
        os.mkdir(img_directory)

    # 이미지를 저장할 폴더
    img_directory = 'C:\\Users\\ooo33.DESKTOP-56U45AS\\Desktop\\img\\{}\\{}'.format(cate, acc_name)
        
    # 폴더가 없으면 새로 생성
    if not os.path.isdir(img_directory):
        os.mkdir(img_directory)
    print(len(images))
    imgArr = []
    try:
        for index, image in enumerate(images):
            link = image.get_attribute('data-src')
            if(link != None):
                start = link.rfind('.')
                end = link.rfind('?')
                filetype = link[start:end]
                name = uuid.uuid1()
                urlretrieve('http:{}'.format(link), 'C:/Users/ooo33.DESKTOP-56U45AS/Desktop/img/{}/{}/{}{}'.format(cate, acc_name, name, '.jpg'))
                imgArr.append('{}{}'.format(name, '.jpg'))
                time.sleep(0.1)
            if(index == 4):
                break
        return "/".join(imgArr)
    except:
        return None

# 판매자 정보를 가공해서 반환
def get_sellerInfo(sellerInfo, n):
    list = {}
    for index, info in enumerate(sellerInfo):
        if "상호" in info:
            list['company']= sellerInfo[index+1]
        if "대표자명" in info:
            list['ceo']= sellerInfo[index+1]
        if "주소" in info:
            list['addr']= sellerInfo[index+1]
        if "전화번호" in info:
            list['tel']= sellerInfo[index+1]
        if "이메일" in info:
            list['userId_id']= str(n) + sellerInfo[index+1] 
        if "사업자번호" in info:
            list['bis_bizRegNum']= sellerInfo[index+1]
        if list.get('userId_id') == None:
            list['userId_id'] = '{}{}@naver.com'.format("b", n+100)

        list['bis_openDate'] = '2010-10-12'
        list['userId_nick'] = get_nick().replace(" ", "") + str(n)
    return list

# 닉네임 자동 생성
def get_nick():
    try:
        url = 'https://nickname.hwanmoo.kr/?format=json&count=1&max_length=5'
        api_json = json.loads(requests.get(url).text)
        nick = api_json['words'][0]
        return nick
    except:
        return str(randrange(100))

# 새로운 엑셀파일 생성
workbook = Workbook()

workbook.create_sheet('숙소')
workbook.create_sheet('객실')

# 첫번째 시트 활성화
sheet = workbook.active
sheet.title = '판매자'


sheet = workbook["판매자"]
sheet.append(['userId_id', 'userId_nick', 'userId_type', 'user_hp', 'bis_company', 'bis_ceo', 'bis_openDate', 'bis_bizRegNum', 'bis_tel', 'bis_zip',
    'bis_addr', 'bis_addrDetail'])

sheet = workbook["숙소"]
sheet.append([
            'acc_id', 'userId_id', 'acc_name', 'accType_no', 'province_no', 'acc_city', 'acc_zip', 
            'acc_addr', 'acc_addrDetail', 'acc_longtitude', 'acc_lattitude', 'acc_info', 
            'acc_comment', 'acc_thumbs', 'acc_checkIn', 'acc_checkOut'
            ])

sheet = workbook["객실"]
sheet.append(['room_id', 'acc_id', 'room_name', 'room_stock', 'room_price', 'room_info', 'room_thumb', 'room_checkIn', 'room_checkOut'])


# 가상 브라우저 실행
chrome_options = Options()
chrome_options.add_experimental_option("detach", True)
browser = webdriver.Chrome('./chromedriver.exe', options=chrome_options)

citys = ['경상북도', '경기도', '충청북도', '충청남도', '강원도', '제주도', '전라북도', '전라남도'
        , '서울', '인천', '광주', '대전', '대구', '울산', '부산']
# citys = ['경상남도']

# 모텔, 호텔, 펜션, 캠핑, 게스트하우스
cates = [1, 2, 3, 5, 6]

acc = []
AccCnt = 344
roomCnt = 2984
for city in citys:
    for cate in cates:
        # 여기어때 검색 페이지 이동
        search = 'https://www.goodchoice.kr/product/result?keyword={}&adcno%5B%5D={}&sel_date=2023-03-30&sel_date2=2023-03-31'.format(city, cate)
        browser.get(search)

        # 검색된 숙소 리스트
        lists = browser.find_elements(By.CSS_SELECTOR, "#poduct_list_area > ul > li")

        # 상세보기 페이지 링크를 저장할 배열 선언
        links = []
        print('갯수',len(lists))

        # 배열에 링크 주소 값 저장
        for list in lists:
            a = list.find_element(By.CSS_SELECTOR , 'a')
            links.append(a.get_attribute('href'))

        # 검색 페이지 닫기
        browser.close

        
        # 각 숙소 리스트 상세 페이지 브라우저 열기
        for index, link in enumerate(links):
            
            # 저장한 링크 주소의 페이지를 연다.
            browser.get(link)

            html = browser.page_source
            soup = bs(html, 'html.parser')
            try:
                # 숙소의 모든 정보를 저장할 변수
                content = {}

                # 숙소의 개별적 정보
                title = browser.find_element(By.CSS_SELECTOR , '#content > div.top > div.right > div.info > h2').text
                addr = browser.find_element(By.CSS_SELECTOR , '#content > div.top > div.right > div.info > p.address').text
                try:
                    comment = browser.find_element(By.CSS_SELECTOR , '#content > div.top > div.right > div.comment > div').text
                except Exception as e:
                    comment = ""

                
                browser.find_element(By.CSS_SELECTOR , '#content > div.tab > button:nth-child(2)').click()
                acc_info = browser.find_element(By.CSS_SELECTOR , '#content > article.detail_info.on > section.default_info').text
                
            
                sellerInfo = browser.find_element(By.CSS_SELECTOR , '#content section.seller_info').get_attribute('textContent')
                images = browser.find_elements(By.CSS_SELECTOR , '#content > div.top > div.left > div.gallery_pc > div > ul > li.swiper-slide > img')
            except:
                continue

            # 판매자 정보 
            sellerInfoList = [v for v in remove_html(sellerInfo).replace('\t', "").splitlines() if v]
            sellerInfoNew = []
            for sif in sellerInfoList:
                sif = sif.strip()
                if(sif != ""):
                    sellerInfoNew.append(sif)
            content['sellerInfo'] = get_sellerInfo(sellerInfoNew, AccCnt)
            
            # 숙소의 정보들 저장
            acc_lattitude, acc_longtitude, acc_zip, acc_addr, acc_addrDetail, acc_city = get_location(addr)
            if(acc_lattitude == None):
                continue
            content['acc_id'] = AccCnt + 1000000
            content['userId_id'] = content.get('sellerInfo').get('userId_id') # 판매자 아이디
            content['acc_name'] = title                           # 숙소명
            content['accType_no'] = get_accType_no(cate)          # 숙소유형
            content['province_no'] = get_province_no(acc_addr)    # 전국8도 및 광역시
            content['acc_city'] = acc_city                        # 군, 구
            content['acc_zip'] = acc_zip                          # 우편번호
            content['acc_addr'] = acc_addr                        # 기본 주소
            content['acc_addrDetail'] = acc_addrDetail            # 상세 주소
            content['acc_longtitude'] = acc_longtitude            # 경도
            content['acc_lattitude'] = acc_lattitude              # 위도
            content['acc_info'] = remove_html(acc_info)           # 기본 정보
            content['comment'] = remove_html(comment)             # 사장님 한마디
            content['acc_checkIn'] = "14:00"                      # 체크인 시간
            content['acc_checkOut'] = "12:00"                     # 체크아웃 시간
            acc_thumbs = get_imgArrStr(images, content.get('province_no'), content.get('acc_id'))  # 썸네일들
            if(acc_thumbs != None):
                content['acc_thumbs'] = acc_thumbs
            else:
                continue

            #판매자 정보 저장
            sheet = workbook["판매자"]
            sheet.append([content.get('sellerInfo').get('userId_id'), content.get('sellerInfo').get('userId_nick'), 1, content.get('sellerInfo').get('tel'),
                            content.get('sellerInfo').get('company'), content.get('sellerInfo').get('ceo'), content.get('sellerInfo').get('bis_openDate'),
                            content.get('sellerInfo').get('bis_bizRegNum'), content.get('sellerInfo').get('tel'), content.get('acc_zip'), content.get('acc_addr'), content.get('acc_addrDetail')])

            #숙소 정보 저장
            sheet = workbook["숙소"]
            sheet.append([
                        content.get('acc_id'), content.get('userId_id'), content.get('acc_name'), content.get('accType_no'), 
                        content.get('province_no'), content.get('acc_city'), content.get('acc_zip'), content.get('acc_addr'),  
                        content.get('acc_addrDetail'), content.get('acc_longtitude'), content.get('acc_lattitude'),
                        content.get('acc_info'), content.get('acc_comment'), content.get('acc_thumbs'), content.get('acc_checkIn'),
                        content.get('acc_checkOut')
                        ])

            # 숙소 객실 정보
            sheet = workbook["객실"]

            roomList = []
            korean = re.compile('[\u3131-\u3163\uac00-\ud7a3]+')
            roomes = soup.select('#product_filter_form > article > div.room')
            for j, room in enumerate(roomes):
                roomInfo = {}
                
                # 객실 번호
                room_id = roomCnt + 1000000
                roomCnt = roomCnt + 1

                # 객실 이름
                room_name = room.select_one("strong.title").text

                # 객실 재고
                room_stock = len(roomes)
                
                # 객실 가격
                if(cate == 1):
                    room_price = room.select_one("div.info > div:nth-child(2) > div > div > p:nth-child(2)").text
                else:
                    room_price = room.select_one("div.info > div > div > div > p:nth-child(2)").text

                room_price = re.sub(korean, '', room_price)
                room_price = room_price.replace(",", "").replace("\n", "")
                if(room_price == None or room_price.replace(" ", "") == ""):
                    room_price = "50000"

                # 객실 이미지
                roomImages = browser.find_elements(By.CSS_SELECTOR, '#product_filter_form > article > div.room:nth-of-type({}) div.owl-item > li > img'.format(j+2))
                
                room_thumb = get_imgArrStr(roomImages, content.get('province_no'), content.get('acc_id'))

                if(room_thumb == None):
                    room_thumb = ""

                roomInfo['room_id'] = room_id
                roomInfo['room_name'] = room_name
                roomInfo['room_stock'] = room_stock
                roomInfo['room_price'] = room_price

                if(cate == 1):
                    roomInfo['room_info'] = "내용 없음"
                else:
                    info = soup.select('body > div.wrap.show > div.pop_useinfo > div.iscroll_cp > div')
                    for n, room_info in enumerate(info):
                        if(j == n):
                            room_info = remove_html(room_info.get_text()).splitlines()
                            roomInfo['room_info'] = '\n'.join([v for v in room_info if v])

                roomInfo['room_thumb'] = room_thumb
                roomInfo['room_checkIn'] = "14:00"
                roomInfo['room_checkOut'] = "12:00" 

                roomList.append(roomInfo)    
                #객실 정보 저장
                sheet.append([
                    roomInfo.get('room_id'), content.get('acc_id'), roomInfo.get('room_name'), roomInfo.get('room_stock'), 
                    roomInfo.get('room_price'), roomInfo.get('room_info'), roomInfo.get('room_thumb'), roomInfo.get('room_checkIn'), roomInfo.get('room_checkOut')
                ])
                
            content['roomList'] = roomList   # 객실 리스트 저장
            

            #숙소들의 변수를 모은 변수에 저장
            acc.append(content)

            browser.close

            AccCnt = AccCnt + 1
            # if(index == 1):
            #     break

workbook.save('C:/Users/ooo33.DESKTOP-56U45AS/Desktop/lemoDB2.xlsx')
workbook.close()